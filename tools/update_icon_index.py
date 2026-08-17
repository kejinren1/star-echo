#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""update_icon_index.py — items sheet 图标序号 0-53 顺序回写（道具图集重建配套）
Excel 为唯一事实源：改 Excel → excel_export.py 重新导出 items.json。
用法: python tools/update_icon_index.py
"""
import sys
from pathlib import Path

from openpyxl import load_workbook

XLSX = Path("D:/30DAYS/docs/GameData.xlsx")
KEY_ROW = 1  # 表头英文 key 在首行（第 2 行为中文标题，第 3 行起数据）
DATA_START = 3


def main():
    wb = load_workbook(XLSX)  # read_only=False
    ws = wb["items"]
    # 定位 icon_index 列
    col = None
    for c in range(1, ws.max_column + 1):
        if ws.cell(row=KEY_ROW, column=c).value == "icon_index":
            col = c
            break
    if col is None:
        print("FAIL: items sheet 无 icon_index 列")
        return 1
    # 数据行数（非空 id）
    n = 0
    for r in range(DATA_START, ws.max_row + 1):
        if ws.cell(row=r, column=1).value:
            n += 1
    for i in range(n):
        ws.cell(row=DATA_START + i, column=col, value=i)
    wb.save(XLSX)
    print(f"OK: items sheet icon_index 列已回写 0..{n-1}（共 {n} 行）")
    return 0


if __name__ == "__main__":
    sys.exit(main())
