#!/usr/bin/env python3
"""Excel → JSON 导出工具（F1.0-2）：校验 + 导出 + 总览 + manifest。

用法：
    python tools/excel_export.py             # 校验通过后导出 data/*.json + 总览 + manifest
    python tools/excel_export.py --check-only  # 只校验（自动化 git 护栏用，不写盘）
    python tools/excel_export.py --overview    # 只刷新 DATA_OVERVIEW.md + 总览 sheet
"""
from __future__ import annotations

import hashlib
import json
import re
import sys
from collections import OrderedDict, Counter
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

from data_schema import (DATA_DIR, MANIFEST_PATH, OVERVIEW_PATH, SHEETS,
                         XLSX_PATH, dumps_json, loads_json, unflatten)

ROOT = Path(__file__).resolve().parent.parent

# ========== 校验清单（单一事实源：与现有 data/*.json 实测一致，2026-08-10 采集） ==========

VALID_RARITY = {"common", "uncommon", "rare", "epic", "legendary"}
VALID_SLOT = {"passive", "relic"}
VALID_ITEM_CATEGORY = {"attack", "defense", "special", "stat"}
VALID_BEHAVIOR = {"aoe_attack", "charge", "chase", "heal", "ranged", "self_heal",
                  "spawn", "stationary", "zigzag"}
VALID_WEAPON_CATEGORY = {"melee", "ranged", "elemental", "engineering"}
VALID_ENEMY_CATEGORY = {"regular", "elite", "boss"}
VALID_REWARD_TYPE = {"attack_percent", "attack_speed_percent", "gold", "heal_percent",
                     "item", "level_up", "luck", "max_hp", "trade", "weapon_upgrade"}
VALID_ROUTE_TYPE = {"add_node", "difficulty", "flag", "reroute", "unlock_node"}
    # player.gd STAT_MAP 已映射键 + CONSUMED_BONUS_KEYS（P0-Bug2 收口后口径）
    # range：STAT_MAP_EXCLUDED 刻意排除键（口径见 player.gd），但属已知键
KNOWN_EFFECT_KEYS = {
    "max_hp", "speed_percent", "armor", "regen", "hp_regen", "dodge_percent",
    "crit_chance_percent", "attack_speed_percent", "melee_attack_speed_percent",
    "damage_percent", "range_percent", "luck", "pickup_range", "life_steal_percent",
    "crit_damage_percent", "damage_taken_percent", "structure_damage_percent",
    "range",
    # CONSUMED_BONUS_KEYS（有消费方）
    "orbit_blade_count", "elemental_damage", "summon_count",
    # 无消费方但已登记（T-050 待 F1 裁决）—— 新增键会触发 WARN
    "attack_speed_per_different_weapon_percent", "auto_turret_per_wave",
    "boss_elite_damage_percent", "burn_duration_percent",
    "damage_reduction_on_hit_percent", "dodge_heal_amount", "dodge_heal_chance",
    "element_duration_percent", "element_reaction_damage_percent", "engineering",
    "fire_damage_percent", "harvesting", "knockback", "melee_damage",
    "miss_chance_percent", "no_weapon_armor_bonus", "ranged_damage", "reaction_heal",
    "shop_weapon_upgrade", "special_enemies_next_wave", "structure_duration_percent",
    "xp_gain_percent",
}


class Report:
    def __init__(self) -> None:
        self.errors: list[str] = []
        self.warns: list[str] = []
        self.notes: list[str] = []

    def err(self, msg: str) -> None:
        self.errors.append(msg)

    def warn(self, msg: str) -> None:
        self.warns.append(msg)

    def ok(self) -> bool:
        return not self.errors


# ========== 读取 ==========

def read_sheet(ws, rep: Report) -> list[dict]:
    """sheet → 行字典列表（第一行表头；空行跳过）"""
    rows: list[dict] = []
    headers = [c.value for c in ws[1]]
    if not headers or not any(headers):
        return rows
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        rec: dict = {}
        for h, v in zip(headers, row):
            if h is None:
                continue
            h = str(h).strip()
            if v is None or (isinstance(v, str) and v.strip() == ""):
                continue  # 空单元格 = 无键（null 语义等价：DataLoader get() 兜底）
            if isinstance(v, str):
                v = v.strip()
            rec[h] = v
        rows.append(rec)
    return rows


def coerce_num(v):
    """数字归一：int 保持 int、float 保持 float、数字字符串转数字（wave 列等）"""
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return v
    if isinstance(v, str):
        s = v.strip()
        if re.fullmatch(r"-?\d+", s):
            return int(s)
        if re.fullmatch(r"-?\d+\.\d+", s):
            return float(s)
    return v


def sheet_names(wb) -> list[str]:
    return [ws.title for ws in wb.worksheets]


# ========== 校验 ==========

def validate(wb, rep: Report) -> dict[str, list[dict]]:
    tables: dict[str, list[dict]] = {}
    actual = sheet_names(wb)
    expected = [s["sheet"] for s in SHEETS.values()]
    for name in expected:
        if name not in actual:
            rep.err(f"sheet 缺失: {name}")
            continue
        tables[name] = read_sheet(wb[name], rep)
    for name in actual:
        if name not in expected and name not in ("说明", "总览"):
            rep.warn(f"未知 sheet（忽略）: {name}")

    # ---- id 唯一性（支持复合主键 unique_with） ----
    for spec in SHEETS.values():
        key = spec["key"]
        if not key:
            continue
        uniq2 = spec.get("unique_with")
        seen: dict = {}
        for i, row in enumerate(tables.get(spec["sheet"], []), 2):
            v = row.get(key)
            if v is None:
                rep.err(f"{spec['sheet']} 第 {i} 行缺少主键 {key}")
                continue
            k = str(v)
            if uniq2:
                k = f"{k}::{row.get(uniq2)}"
            if k in seen:
                rep.err(f"{spec['sheet']} 主键重复: {v}"
                        + (f" + {uniq2}={row.get(uniq2)}" if uniq2 else "")
                        + f"（第 {seen[k]} 行与第 {i} 行）")
            seen[k] = i

    # ---- 子表引用（父键存在性） ----
    for spec in SHEETS.values():
        if spec["kind"] not in ("child_list", "child_dict"):
            continue
        parent_spec = SHEETS[spec["parent_sheet"]]
        parent_keys = {str(r.get(parent_spec["key"])) for r in tables.get(parent_spec["sheet"], [])}
        for i, row in enumerate(tables.get(spec["sheet"], []), 2):
            fk = str(row.get(spec["key"], ""))
            if fk and fk not in parent_keys:
                rep.err(f"{spec['sheet']} 第 {i} 行引用了不存在的 {parent_spec['sheet']}: {fk}")

    # ---- 枚举 ----
    for i, row in enumerate(tables.get("items", []), 2):
        if row.get("rarity") not in VALID_RARITY:
            rep.err(f"items 第 {i} 行 rarity 非法: {row.get('rarity')}（合法: {sorted(VALID_RARITY)}）")
        if row.get("slot") and row.get("slot") not in VALID_SLOT:
            rep.err(f"items 第 {i} 行 slot 非法: {row.get('slot')}")
        if row.get("category") and row.get("category") not in VALID_ITEM_CATEGORY:
            rep.err(f"items 第 {i} 行 category 非法: {row.get('category')}")
    for i, row in enumerate(tables.get("enemies", []), 2):
        if row.get("behavior") and row.get("behavior") not in VALID_BEHAVIOR:
            rep.err(f"enemies 第 {i} 行 behavior 非法: {row.get('behavior')}")
        if row.get("_xlsx_category") and row.get("_xlsx_category") not in VALID_ENEMY_CATEGORY:
            rep.err(f"enemies 第 {i} 行 _xlsx_category 非法: {row.get('_xlsx_category')}")
    for i, row in enumerate(tables.get("weapons", []), 2):
        if row.get("_xlsx_category") and row.get("_xlsx_category") not in VALID_WEAPON_CATEGORY:
            rep.err(f"weapons 第 {i} 行 _xlsx_category 非法: {row.get('_xlsx_category')}")

    # ---- 引用完整性 ----
    weapon_ids = {str(r.get("id")) for r in tables.get("weapons", []) if r.get("id")}
    for i, row in enumerate(tables.get("characters", []), 2):
        sw = row.get("starting_weapon")
        if sw and str(sw) not in weapon_ids:
            rep.err(f"characters 第 {i} 行 starting_weapon 不存在: {sw}")

    # ---- items_effects 键白名单 ----
    for i, row in enumerate(tables.get("items_effects", []), 2):
        k = row.get("key")
        if k and str(k) not in KNOWN_EFFECT_KEYS:
            rep.warn(f"items_effects 第 {i} 行新效果键（未登记，请补 code 映射）: {k}")

    # ---- 事件奖励类型 ----
    for i, row in enumerate(tables.get("events", []), 2):
        for col in ("choiceA.reward.type", "choiceB.effect_on_route.type"):
            v = row.get(col)
            if v is not None:
                enum = VALID_REWARD_TYPE if ".reward." in col else VALID_ROUTE_TYPE
                if str(v) not in enum:
                    rep.err(f"events 第 {i} 行 {col} 非法: {v}")

    # ---- JSON 列可解析（双形态列豁免：标量 or JSON 文本二选一合法） ----
    DUAL_FORM_COLS = {"choiceA.reward.value", "choiceB.effect_on_route.value"}
    for spec in SHEETS.values():
        jcols = spec["json_cols"]
        if not jcols:
            continue
        for i, row in enumerate(tables.get(spec["sheet"], []), 2):
            for col in jcols:
                if col in DUAL_FORM_COLS:
                    continue
                v = row.get(col)
                if v is not None and loads_json(v) is None:
                    rep.err(f"{spec['sheet']} 第 {i} 行 {col} JSON 解析失败")

    return tables


# ========== 导出 ==========

def build_json_files(tables: dict[str, list[dict]], rep: Report) -> dict[str, object]:
    """按 schema 重建各 JSON 文件内容（结构与原 data/*.json 完全一致）"""
    files: dict[str, object] = {}

    def main_rows(sheet: str) -> list[dict]:
        rows = []
        for r in tables.get(sheet, []):
            rec = {k: coerce_num(v) for k, v in r.items() if not k.startswith("_")}
            rows.append(unflatten(rec, set(SHEETS[sheet]["json_cols"])))
        return rows

    # weapons（category_map + levels 子表回填）
    wrows = []
    for r in tables.get("weapons", []):
        rec = {k: coerce_num(v) for k, v in r.items() if not k.startswith("_")}
        rec = unflatten(rec, set(SHEETS["weapons"]["json_cols"]))
        rec.pop("levels", None)
        rec["levels"] = []
        wrows.append((str(r.get("_xlsx_category", "melee")), rec))
    lv_index: dict[str, list] = {}
    for r in tables.get("weapons_levels", []):
        wid = str(r.get("weapon_id", ""))
        lv = {k: coerce_num(v) for k, v in r.items() if k != "weapon_id"}
        lv = unflatten(lv, set())
        lv_index.setdefault(wid, []).append(lv)
    for cat, rec in wrows:
        rec["levels"] = lv_index.get(str(rec.get("id")), [])
    weapons_out: dict = {}
    for cat, rec in wrows:
        weapons_out.setdefault(cat, []).append(rec)
    files["weapons.json"] = {"weapons": weapons_out}

    # items（list + effects 子表回填）
    irows = []
    for r in tables.get("items", []):
        rec = {k: coerce_num(v) for k, v in r.items() if not k.startswith("_")}
        rec = unflatten(rec, set(SHEETS["items"]["json_cols"]))
        rec.pop("effects", None)
        rec["effects"] = {}
        irows.append(rec)
    eff_index: dict[str, dict] = {}
    for r in tables.get("items_effects", []):
        eff_index.setdefault(str(r.get("item_id", "")), {})[str(r.get("key"))] = coerce_num(r.get("value"))
    for rec in irows:
        rec["effects"] = eff_index.get(str(rec.get("id")), {})
    files["items.json"] = {"items": irows}

    # enemies（category_map + scaling）
    erows: dict = {}
    for r in tables.get("enemies", []):
        rec = {k: coerce_num(v) for k, v in r.items() if not k.startswith("_")}
        rec = unflatten(rec, set(SHEETS["enemies"]["json_cols"]))
        erows.setdefault(str(r.get("_xlsx_category", "regular")), []).append(rec)
    scaling_rows = tables.get("enemy_scaling", [])
    scaling = {k: coerce_num(v) for k, v in (scaling_rows[0].items() if scaling_rows else {}.items())}
    files["enemies.json"] = {"enemies": erows, "scaling": scaling}

    # characters（list + skill 点号 + passives/penalties 回填）
    crows = main_rows("characters")
    for rec in crows:
        rec.pop("passive", None)
        rec.pop("penalty", None)
    for sheet, target in (("characters_passives", "passive"), ("characters_penalties", "penalty")):
        idx: dict[str, dict] = {}
        for r in tables.get(sheet, []):
            idx.setdefault(str(r.get("char_id", "")), {})[str(r.get("key"))] = coerce_num(r.get("value"))
        for rec in crows:
            data = idx.get(str(rec.get("id")))
            if data:  # 无数据不写空字典（与原 JSON 一致：部分角色无 penalty 键）
                rec[target] = data
    files["characters.json"] = {"characters": crows}

    # waves（list + generation/rewards）
    wrows2 = []
    for r in tables.get("waves", []):
        rec = {k: coerce_num(v) for k, v in r.items() if not k.startswith("_")}
        rec = unflatten(rec, set(SHEETS["waves"]["json_cols"]))
        wrows2.append(rec)
    def flat_sheet(sheet: str) -> dict:
        rows = tables.get(sheet, [])
        if not rows:
            return {}
        out: dict = {}
        for k, v in rows[0].items():
            if k.startswith("_"):
                continue
            out[k] = coerce_num(v)
        return unflatten(out, set(SHEETS[sheet]["json_cols"]))
    files["waves.json"] = {
        "waves": wrows2,
        "generation": flat_sheet("wave_generation"),
        "rewards": flat_sheet("wave_rewards"),
    }

    # events（list + choice 点号展开；value 双形态列经 json_cols 解析）
    erows2 = []
    for r in tables.get("events", []):
        rec = {k: coerce_num(v) for k, v in r.items() if not k.startswith("_")}
        rec = unflatten(rec, set(SHEETS["events"]["json_cols"]))
        erows2.append(rec)
    files["events.json"] = {"events": erows2}

    # stats（category_map + formulas/leveling）
    srows: dict = {}
    for r in tables.get("stats", []):
        rec = {k: coerce_num(v) for k, v in r.items() if not k.startswith("_")}
        rec = unflatten(rec, set())
        srows.setdefault(str(r.get("_xlsx_category", "base")), []).append(rec)
    files["stats.json"] = {
        "stats": srows,
        "formulas": flat_sheet("stats_formulas"),
        "leveling": flat_sheet("stats_leveling"),
        "shop": flat_sheet("stats_shop"),
    }

    # elements（dict + reactions + reaction_rules）
    el_rows = tables.get("elements", [])
    els: dict = {}
    for r in el_rows:
        eid = str(r.get("element_id"))
        rec = {k: coerce_num(v) for k, v in r.items() if k != "element_id" and not k.startswith("_")}
        els[eid] = unflatten(rec, set())
    react_rows = []
    for r in tables.get("element_reactions", []):
        rec = {k: coerce_num(v) for k, v in r.items() if k != "index" and not k.startswith("_")}
        react_rows.append(unflatten(rec, set(SHEETS["element_reactions"]["json_cols"])))
    files["elements.json"] = {"elemental_status": els, "element_reactions": react_rows}
    rr = flat_sheet("reaction_rules")
    if rr:
        files["elements.json"]["reaction_rules"] = rr

    # routes（flat_dict 单行）
    files["routes.json"] = flat_sheet("routes")
    return files


# ========== 总览 ==========

def build_overview(tables: dict[str, list[dict]], rep: Report) -> str:
    lines: list[str] = []
    lines.append("# 数据分布总览（DATA_OVERVIEW）")
    lines.append("")
    lines.append(f"- 生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M')}")
    lines.append("- 来源：docs/GameData.xlsx（tools/excel_export.py 导出时自动刷新）")
    lines.append("")

    order = ["weapons", "items", "enemies", "characters", "waves", "events",
             "stats", "elements", "element_reactions", "routes"]
    for sheet in order:
        rows = tables.get(sheet, [])
        if not rows:
            continue
        lines.append(f"## {sheet}（{len(rows)} 行）")
        if sheet == "items":
            cnt = Counter(str(r.get("rarity", "?")) for r in rows)
            lines.append(f"- 稀有度分布：{dict(cnt)}")
            effs = tables.get("items_effects", [])
            kc = Counter(str(r.get("key")) for r in effs)
            top = ", ".join(f"{k}×{n}" for k, n in kc.most_common(8))
            lines.append(f"- 效果键 Top（共 {len(kc)} 种）：{top}")
            # 数值范围抽样
            for stat in ("price",):
                vals = [float(r[stat]) for r in rows if r.get(stat) is not None]
                if vals:
                    lines.append(f"- {stat}：min {min(vals)} / max {max(vals)} / 均值 {sum(vals)/len(vals):.1f}")
        elif sheet == "weapons":
            cnt = Counter(str(r.get("_xlsx_category", "?")) for r in rows)
            lines.append(f"- 分类分布：{dict(cnt)}")
            lv = tables.get("weapons_levels", [])
            lines.append(f"- 等级条目：{len(lv)}（{len(lv)//8 if lv else 0} 把 × 8 级）")
            for stat in ("damage", "cooldown", "range"):
                vals = [float(r[stat]) for r in rows if r.get(stat) is not None]
                if vals:
                    lines.append(f"- 基础 {stat}：min {min(vals)} / max {max(vals)} / 均值 {sum(vals)/len(vals):.1f}")
        elif sheet == "enemies":
            cnt = Counter(str(r.get("_xlsx_category", "?")) for r in rows)
            lines.append(f"- 分类分布：{dict(cnt)}")
        elif sheet == "characters":
            sw = [str(r.get("starting_weapon")) for r in rows if r.get("starting_weapon")]
            lines.append(f"- 起始武器 {len(sw)} 把：{', '.join(sw)}")
        elif sheet == "waves":
            lines.append(f"- 波次范围：{min(int(r['wave']) for r in rows if r.get('wave'))} – "
                         f"{max(int(r['wave']) for r in rows if r.get('wave'))}")
        lines.append("")

    # 未消费/死数据提示
    lines.append("## 关注项")
    unmapped = [k for k in KNOWN_EFFECT_KEYS if k not in {
        "max_hp", "speed_percent", "armor", "regen", "hp_regen", "dodge_percent",
        "crit_chance_percent", "attack_speed_percent", "melee_attack_speed_percent",
        "damage_percent", "range_percent", "luck", "pickup_range", "life_steal_percent",
        "crit_damage_percent", "damage_taken_percent", "structure_damage_percent",
        "orbit_blade_count", "elemental_damage", "summon_count",
    }]
    lines.append(f"- 无消费方效果键 {len(unmapped)} 个（T-050，待 F1 逐键裁决）：{', '.join(sorted(unmapped))}")
    lines.append("")
    return "\n".join(lines)


def refresh_overview_sheet(wb, overview_text: str) -> None:
    if "总览" in wb.sheetnames:
        del wb["总览"]
    ws = wb.create_sheet("总览")
    ws.sheet_properties.tabColor = "808080"
    for i, line in enumerate(overview_text.splitlines(), 1):
        ws.cell(row=i, column=1, value=line)
    ws.column_dimensions["A"].width = 120


def equivalent(a, b) -> bool:
    """规范化等价：null 与「键缺失」视为相同（DataLoader get() 兜底语义）"""
    if isinstance(a, dict) and isinstance(b, dict):
        keys = set(a) | set(b)
        return all(equivalent(a.get(k), b.get(k)) for k in keys)
    if isinstance(a, list) and isinstance(b, list):
        return len(a) == len(b) and all(equivalent(x, y) for x, y in zip(a, b))
    if a is None or b is None:
        return a is None and b is None
    return a == b


# ========== 主流程 ==========

def main() -> int:
    check_only = "--check-only" in sys.argv
    only_overview = "--overview" in sys.argv

    rep = Report()
    if not XLSX_PATH.exists():
        rep.err(f"工作簿不存在: {XLSX_PATH}（先运行 tools/json_to_excel.py）")
        print(rep.errors)
        return 1

    wb = load_workbook(XLSX_PATH)
    tables = validate(wb, rep)

    if rep.errors:
        print("=== 校验失败（不导出）===")
        for e in rep.errors:
            print("  ERR ", e)
        for w in rep.warns:
            print("  WARN", w)
        return 1

    if only_overview:
        text = build_overview(tables, rep)
        refresh_overview_sheet(wb, text)
        wb.save(XLSX_PATH)
        OVERVIEW_PATH.write_text(text, encoding="utf-8")
        print(f"[export] 总览已刷新 → {OVERVIEW_PATH} / 总览 sheet")
        return 0

    files = build_json_files(tables, rep)

    # roundtrip 自检：导出结果与原 JSON 规范化对比（null 与缺失等价；首次导入需人工确认 diff）
    for name, content in files.items():
        path = DATA_DIR / name
        if path.exists():
            old = json.loads(path.read_text(encoding="utf-8"))
            if not equivalent(old, content):
                rep.warn(f"导出与现有 {name} 有差异（首次导入需人工确认；diff 见 tools/export_diff 对比）")
        (DATA_DIR / name).write_text(
            json.dumps(content, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    # manifest 指纹
    fp_parts = []
    for name in sorted(files):
        raw = (DATA_DIR / name).read_bytes()
        fp_parts.append(f"{name}:{hashlib.sha256(raw).hexdigest()[:12]}")
    manifest = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "fingerprint": hashlib.sha256("|".join(fp_parts).encode()).hexdigest()[:16],
        "files": {name: fp.split(":")[1] for fp in fp_parts},
    }
    MANIFEST_PATH.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")

    # 总览
    text = build_overview(tables, rep)
    refresh_overview_sheet(wb, text)
    wb.save(XLSX_PATH)
    OVERVIEW_PATH.write_text(text, encoding="utf-8")

    print("=== 导出完成 ===")
    for name in files:
        print(f"  [out] data/{name}")
    print(f"  [out] {MANIFEST_PATH}")
    print(f"  [out] {OVERVIEW_PATH} + 总览 sheet")
    for w in rep.warns:
        print("  WARN", w)
    return 0


if __name__ == "__main__":
    sys.exit(main())
