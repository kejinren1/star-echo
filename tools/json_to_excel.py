#!/usr/bin/env python3
"""JSON → Excel 导入工具（F1.0-1）：现有 data/*.json 一次性生成 docs/GameData.xlsx。

用法：
    python tools/json_to_excel.py
输出：
    docs/GameData.xlsx（说明 sheet + 各数据表 sheet；嵌套拆子表，规则见 data_schema.py）
"""
from __future__ import annotations

import json
import sys
from collections import OrderedDict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from data_schema import SHEETS, XLSX_PATH, dumps_json, flatten

ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = ROOT / "data"

HEADER_FILL = PatternFill("solid", fgColor="DDEBF7")
KEY_FILL = PatternFill("solid", fgColor="FFF2CC")
NOTE_FONT = Font(size=10, color="555555")


def load_json(name: str):
    with open(DATA_DIR / name, encoding="utf-8") as f:
        return json.load(f)


def collect_columns(rows: list[dict], json_cols: set, excluded_prefix: str = "") -> list[str]:
    """行集合的列并集（保持发现顺序，主键排最前）"""
    cols: "OrderedDict[str, None]" = OrderedDict()
    for row in rows:
        for k in row.keys():
            if excluded_prefix and k.startswith(excluded_prefix):
                continue
            if k not in cols:
                cols[k] = None
    return list(cols)


def main() -> int:
    wb = Workbook()
    wb.remove(wb.active)

    # ===== 说明 sheet =====
    readme = wb.create_sheet("说明")
    readme["A1"] = "《星骸回响》游戏数据工作簿（docs/GameData.xlsx）"
    readme["A1"].font = Font(size=14, bold=True)
    readme.append([])
    for line in [
        "【使用流程】改本工作簿 → 运行 python tools/excel_export.py → 校验通过自动导出 data/*.json → 游戏生效",
        "【唯一事实源】本工作簿是唯一编辑入口；data/*.json 为导出产物（generated），禁止手改",
        "【表结构】主表一行一实体；嵌套数据拆子表（weapons_levels / items_effects / characters_passives 等）",
        "【点号列】形如 scaling.melee_damage / skill.effects.shield 的列 = 嵌套对象字段，直接改数值即可",
        "【JSON 列】evolution / composition / phases 等列内容为 JSON 文本，仅进阶调整时编辑",
        "【分类列】weapons / enemies / stats 表带 _xlsx_category 列（导出时用于归组，勿改）",
        "【校验】导出前自动检查：必填/唯一/类型/枚举/引用完整性；失败即中断，按提示修改后重跑",
        "【总览】导出时刷新『总览』sheet 与 docs/DATA_OVERVIEW.md，可随时查看数据分布",
        "【铁律】数值/描述/条件参数进表；行为逻辑永远在代码（不要把 if 逻辑写进表格）",
    ]:
        readme.append([line])
    readme.column_dimensions["A"].width = 130
    for r in range(3, readme.max_row + 1):
        readme.cell(row=r, column=1).font = NOTE_FONT

    # ===== 数据表 =====
    loaded_files: dict[str, dict] = {}

    def file_data(name: str) -> dict:
        if name not in loaded_files:
            loaded_files[name] = load_json(name)
        return loaded_files[name]

    # 先收集所有主表行（便于算子表数据），按 SHEETS 顺序处理
    for spec in SHEETS.values():
        if spec["kind"] in ("child_list", "child_dict"):
            continue
        data = file_data(spec["file"])
        sheet_name = spec["sheet"]
        ws = wb.create_sheet(sheet_name)
        json_cols = set(spec["json_cols"])
        dict_sub_cols = set(spec.get("dict_sub_cols", []))

        # ---- 提取行（统一为 list[dict]） ----
        rows: list[dict] = []
        if spec["kind"] == "category_map":
            for cat, items in data.get(spec["root"], {}).items():
                for it in items:
                    it = dict(it)
                    it["_xlsx_category"] = cat
                    rows.append(it)
        elif spec["kind"] == "list":
            rows = [dict(it) for it in data.get(spec["root"], [])]
            if spec.get("auto_key"):
                for i, it in enumerate(rows):
                    it.setdefault(spec["key"], i + 1)
        elif spec["kind"] == "dict":
            rows = []
            for kid, it in data.get(spec["root"], {}).items():
                it = dict(it)
                it[spec["key"]] = kid
                rows.append(it)
        elif spec["kind"] == "flat_dict":
            src = data if spec["root"] is None else data.get(spec["root"], {})
            if src:
                flat_row: dict = {}
                flatten("", dict(src), set(spec["json_cols"]), flat_row)
                rows = [flat_row]

        # ---- 拆子表数据 ----
        child_specs = [SHEETS[c] for c in (spec["child"],) if spec.get("child")]
        child_specs += [SHEETS[c] for c in (spec.get("dict_sub_cols") or [])]
        child_rows: dict[str, list[dict]] = {c["sheet"]: [] for c in child_specs}
        for row in rows:
            # dict 子表：passive/penalty（键值两列）
            for cspec in child_specs:
                if cspec["kind"] == "child_dict":
                    sub_key = cspec["key"]
                    parent_key = spec["key"]
                    for k, v in row.pop(cspec["root"], {}).items():
                        child_rows[cspec["sheet"]].append({sub_key: row[parent_key], "key": k, "value": v})
                elif cspec["kind"] == "child_list":
                    sub_key = cspec["key"]
                    parent_key = spec["key"]
                    lst = row.pop(cspec["root"], [])
                    for i, it in enumerate(lst):
                        it = dict(it)
                        it[sub_key] = row[parent_key]
                        if cspec["sheet"] == "weapons_levels" and "level" not in it:
                            it["level"] = i + 1
                        child_rows[cspec["sheet"]].append(it)
            # dict 点号列展开（skill / scaling 等：skill.cooldown、skill.effects.shield 直接可改）
            for k in list(row.keys()):
                v = row[k]
                if isinstance(v, dict) and k not in json_cols and k not in dict_sub_cols:
                    flat: dict = {}
                    flatten(k, v, set(json_cols), flat)  # 必须透传 json_cols：value 双形态列在递归时保持 JSON 文本
                    row.pop(k)
                    row.update(flat)
                elif isinstance(v, (list, dict)) and k in json_cols:
                    row[k] = dumps_json(v)

        # ---- 写主表 ----
        cols = collect_columns(rows, json_cols)
        for ci, col in enumerate(cols, 1):
            cell = ws.cell(row=1, column=ci, value=col)
            cell.fill = KEY_FILL if col == spec["key"] else HEADER_FILL
            cell.font = Font(bold=True, size=10)
        for ri, row in enumerate(rows, 2):
            for ci, col in enumerate(cols, 1):
                v = row.get(col)
                if isinstance(v, (dict, list)):
                    v = dumps_json(v)
                ws.cell(row=ri, column=ci, value=v)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"
        for ci, col in enumerate(cols, 1):
            width = min(max(len(str(col)) + 2, 8), 40)
            ws.column_dimensions[get_column_letter(ci)].width = width
        print(f"[import] {sheet_name}: {len(rows)} 行 × {len(cols)} 列")

        # ---- 写子表 ----
        for cspec in child_specs:
            crows = child_rows[cspec["sheet"]]
            cws = wb.create_sheet(cspec["sheet"])
            ccols = collect_columns(crows, set(cspec["json_cols"]))
            for ci, col in enumerate(ccols, 1):
                cell = cws.cell(row=1, column=ci, value=col)
                cell.fill = KEY_FILL if col == cspec["key"] else HEADER_FILL
                cell.font = Font(bold=True, size=10)
            for ri, row in enumerate(crows, 2):
                for ci, col in enumerate(ccols, 1):
                    v = row.get(col)
                    if isinstance(v, (dict, list)):
                        v = dumps_json(v)
                    cws.cell(row=ri, column=ci, value=v)
            cws.freeze_panes = "A2"
            cws.auto_filter.ref = f"A1:{get_column_letter(len(ccols))}1"
            for ci, col in enumerate(ccols, 1):
                cws.column_dimensions[get_column_letter(ci)].width = min(max(len(str(col)) + 2, 8), 30)
            print(f"[import] {cspec['sheet']}: {len(crows)} 行 × {len(ccols)} 列")

    wb.save(XLSX_PATH)
    print(f"[import] 完成 → {XLSX_PATH}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
